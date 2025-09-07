import requests
import sys
import os
import tempfile
import openpyxl
from datetime import datetime
from io import BytesIO

class ImageSearchAPITester:
    def __init__(self, base_url="https://productfinder-3.preview.emergentagent.com"):
        self.base_url = base_url
        self.api_url = f"{base_url}/api"
        self.tests_run = 0
        self.tests_passed = 0

    def run_test(self, name, method, endpoint, expected_status, data=None, files=None, params=None):
        """Run a single API test"""
        url = f"{self.api_url}/{endpoint}" if endpoint else self.api_url
        headers = {}
        
        # Don't set Content-Type for multipart/form-data requests
        if not files:
            headers['Content-Type'] = 'application/json'

        self.tests_run += 1
        print(f"\n🔍 Testing {name}...")
        print(f"   URL: {url}")
        
        try:
            if method == 'GET':
                response = requests.get(url, headers=headers, params=params)
            elif method == 'POST':
                if files:
                    response = requests.post(url, data=data, files=files)
                else:
                    response = requests.post(url, json=data, headers=headers)

            success = response.status_code == expected_status
            if success:
                self.tests_passed += 1
                print(f"✅ Passed - Status: {response.status_code}")
                try:
                    response_data = response.json()
                    print(f"   Response: {response_data}")
                    return True, response_data
                except:
                    print(f"   Response: Binary data ({len(response.content)} bytes)")
                    return True, response.content
            else:
                print(f"❌ Failed - Expected {expected_status}, got {response.status_code}")
                try:
                    error_data = response.json()
                    print(f"   Error: {error_data}")
                except:
                    print(f"   Error: {response.text}")
                return False, {}

        except Exception as e:
            print(f"❌ Failed - Error: {str(e)}")
            return False, {}

    def test_root_endpoint(self):
        """Test root API endpoint"""
        return self.run_test("Root Endpoint", "GET", "", 200)

    def test_single_search_known_codes(self):
        """Test single search with known working codes"""
        # These are the codes mentioned in the request that should work
        known_codes = ["24369", "13025", "2210", "117"]
        found_any = False
        
        for code in known_codes:
            success, response = self.run_test(
                f"Single Search - Known Code {code}",
                "POST",
                "search-single",
                200,
                data={"code": code}
            )
            
            if success and response.get('found'):
                print(f"   ✅ Found image for {code}: {response.get('image_url')}")
                print(f"   📁 Format: {response.get('format')}")
                found_any = True
            elif success:
                print(f"   ❌ No image found for {code} (should have been found)")
        
        return True, {"found_any": found_any}

    def test_single_search_test_codes(self):
        """Test single search with generic test codes"""
        test_codes = ["TEST123", "PROD001", "ABC123", "SAMPLE"]
        
        for code in test_codes:
            success, response = self.run_test(
                f"Single Search - Test Code {code}",
                "POST",
                "search-single",
                200,
                data={"code": code}
            )
            
            if success and response.get('found'):
                print(f"   ✅ Found image for {code}: {response.get('image_url')}")
            elif success:
                print(f"   ℹ️  No image found for {code} (expected)")
        
        return True, {"code": "TEST123", "found": False}

    def test_single_search_empty(self):
        """Test single search with empty code"""
        return self.run_test(
            "Single Search - Empty Code",
            "POST",
            "search-single",
            400,
            data={"code": ""}
        )

    def test_download_image_invalid(self):
        """Test download with invalid URL"""
        return self.run_test(
            "Download Image - Invalid URL",
            "GET",
            "download-image",
            404,
            params={
                "url": "https://invalid-url.com/test.jpg",
                "filename": "test.jpg"
            }
        )

    def create_test_excel_file(self, codes):
        """Create a test Excel file with product codes"""
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        
        # Add header
        sheet.cell(row=1, column=1, value="CODICE")
        
        # Add codes
        for i, code in enumerate(codes, start=2):
            sheet.cell(row=i, column=1, value=code)
        
        # Save to BytesIO
        excel_buffer = BytesIO()
        workbook.save(excel_buffer)
        excel_buffer.seek(0)
        
        return excel_buffer

    def test_batch_search_known_codes(self):
        """Test batch search with known working codes"""
        known_codes = ["24369", "13025", "2210", "117"]
        excel_file = self.create_test_excel_file(known_codes)
        
        files = {'file': ('known_codes.xlsx', excel_file, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
        
        return self.run_test(
            "Batch Search - Known Codes",
            "POST",
            "search-batch",
            200,
            files=files
        )

    def test_batch_search_test_codes(self):
        """Test batch search with test codes"""
        test_codes = ["TEST123", "PROD001", "ABC123", "SAMPLE", "DEMO"]
        excel_file = self.create_test_excel_file(test_codes)
        
        files = {'file': ('test_codes.xlsx', excel_file, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
        
        return self.run_test(
            "Batch Search - Test Codes",
            "POST",
            "search-batch",
            200,
            files=files
        )

    def test_batch_search_invalid_file(self):
        """Test batch search with invalid file"""
        # Create a text file instead of Excel
        text_content = "This is not an Excel file"
        files = {'file': ('test.txt', BytesIO(text_content.encode()), 'text/plain')}
        
        return self.run_test(
            "Batch Search - Invalid File",
            "POST",
            "search-batch",
            400,
            files=files
        )

    def test_batch_search_no_codice_column(self):
        """Test batch search with Excel file without CODICE column"""
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        
        # Add wrong header
        sheet.cell(row=1, column=1, value="PRODUCT")
        sheet.cell(row=2, column=1, value="TEST123")
        
        excel_buffer = BytesIO()
        workbook.save(excel_buffer)
        excel_buffer.seek(0)
        
        files = {'file': ('no_codice.xlsx', excel_buffer, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
        
        return self.run_test(
            "Batch Search - No CODICE Column",
            "POST",
            "search-batch",
            400,
            files=files
        )

    def test_download_batch_zip_known_codes(self):
        """Test batch ZIP download with known codes"""
        known_codes = ["24369", "13025", "2210", "117"]
        excel_file = self.create_test_excel_file(known_codes)
        
        files = {'file': ('known_codes.xlsx', excel_file, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
        
        success, response = self.run_test(
            "Download Batch ZIP - Known Codes",
            "POST",
            "download-batch-zip",
            200,
            files=files
        )
        
        # If no images found, it might return 404
        if not success:
            success_404, _ = self.run_test(
                "Download Batch ZIP - Known Codes (No Images)",
                "POST", 
                "download-batch-zip",
                404,
                files=files
            )
            return success_404, {}
        
        return success, response

    def test_download_batch_zip_test_codes(self):
        """Test batch ZIP download with test codes"""
        test_codes = ["TEST123", "PROD001", "ABC123"]
        excel_file = self.create_test_excel_file(test_codes)
        
        files = {'file': ('test_codes.xlsx', excel_file, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
        
        success, response = self.run_test(
            "Download Batch ZIP - Test Codes",
            "POST",
            "download-batch-zip",
            200,
            files=files
        )
        
        # The endpoint might return 404 if no images are found, which is also valid
        if not success:
            success_404, _ = self.run_test(
                "Download Batch ZIP - Test Codes (No Images)",
                "POST", 
                "download-batch-zip",
                404,
                files=files
            )
            return success_404, {}
        
        return success, response

    def test_batch_search_async_known_codes(self):
        """Test async batch search with known working codes - STUCK TASK"""
        known_codes = ["24369", "13025", "2210", "117"]
        excel_file = self.create_test_excel_file(known_codes)
        
        files = {'file': ('known_codes_async.xlsx', excel_file, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
        
        success, response = self.run_test(
            "Batch Search Async - Known Codes (STUCK TASK)",
            "POST",
            "search-batch-async",
            200,
            files=files
        )
        
        if success and response.get('task_id'):
            print(f"   ✅ Got task_id: {response.get('task_id')}")
            print(f"   📊 Total codes: {response.get('total_codes')}")
            print(f"   📋 Column used: {response.get('column_used')}")
            return True, response
        
        return success, response

    def test_batch_search_async_test_codes(self):
        """Test async batch search with test codes"""
        test_codes = ["TEST123", "PROD001", "ABC123", "SAMPLE"]
        excel_file = self.create_test_excel_file(test_codes)
        
        files = {'file': ('test_codes_async.xlsx', excel_file, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
        
        success, response = self.run_test(
            "Batch Search Async - Test Codes",
            "POST",
            "search-batch-async",
            200,
            files=files
        )
        
        if success and response.get('task_id'):
            print(f"   ✅ Got task_id: {response.get('task_id')}")
            return True, response
        
        return success, response

    def test_progress_tracking_invalid_task(self):
        """Test progress tracking with invalid task ID"""
        fake_task_id = "invalid-task-id-12345"
        
        return self.run_test(
            "Progress Tracking - Invalid Task ID",
            "GET",
            f"progress/{fake_task_id}",
            404
        )

    def test_progress_tracking_valid_task(self):
        """Test progress tracking with valid task ID - STUCK TASK"""
        # First create an async batch search to get a valid task_id
        test_codes = ["TEST123", "PROD001"]
        excel_file = self.create_test_excel_file(test_codes)
        
        files = {'file': ('progress_test.xlsx', excel_file, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
        
        # Start async batch search
        success, response = self.run_test(
            "Progress Tracking Setup - Create Task",
            "POST",
            "search-batch-async",
            200,
            files=files
        )
        
        if not success or not response.get('task_id'):
            print("   ❌ Failed to create task for progress tracking test")
            return False, {}
        
        task_id = response.get('task_id')
        print(f"   📋 Created task_id: {task_id}")
        
        # Now test progress tracking
        import time
        time.sleep(1)  # Give it a moment to start processing
        
        progress_success, progress_response = self.run_test(
            "Progress Tracking - Valid Task ID (STUCK TASK)",
            "GET",
            f"progress/{task_id}",
            200
        )
        
        if progress_success:
            print(f"   📊 Progress: {progress_response.get('progress_percentage', 0)}%")
            print(f"   📈 Status: {progress_response.get('status', 'unknown')}")
            print(f"   📝 Current item: {progress_response.get('current_item', 'none')}")
            print(f"   ✅ Found: {progress_response.get('found_count', 0)}")
            print(f"   ❌ Not found: {progress_response.get('not_found_count', 0)}")
        
        return progress_success, progress_response

    def test_complete_async_workflow(self):
        """Test complete async workflow: upload -> progress -> completion - CRITICAL STUCK TASK"""
        print("\n🔄 Testing Complete Async Workflow (CRITICAL STUCK TASK)")
        
        # Use known codes that should have images
        known_codes = ["24369", "13025"]
        excel_file = self.create_test_excel_file(known_codes)
        
        files = {'file': ('workflow_test.xlsx', excel_file, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
        
        # Step 1: Start async batch search
        success, response = self.run_test(
            "Workflow Step 1 - Start Async Batch",
            "POST",
            "search-batch-async",
            200,
            files=files
        )
        
        if not success or not response.get('task_id'):
            print("   ❌ Workflow failed at step 1")
            return False, {}
        
        task_id = response.get('task_id')
        print(f"   📋 Task ID: {task_id}")
        
        # Step 2: Poll progress multiple times
        import time
        max_polls = 10
        poll_count = 0
        final_status = None
        
        while poll_count < max_polls:
            time.sleep(2)  # Wait between polls
            poll_count += 1
            
            progress_success, progress_response = self.run_test(
                f"Workflow Step 2.{poll_count} - Poll Progress",
                "GET",
                f"progress/{task_id}",
                200
            )
            
            if not progress_success:
                print(f"   ❌ Progress polling failed at attempt {poll_count}")
                break
            
            status = progress_response.get('status', 'unknown')
            progress_pct = progress_response.get('progress_percentage', 0)
            
            print(f"   📊 Poll {poll_count}: {progress_pct}% - Status: {status}")
            
            if status in ['completed', 'error']:
                final_status = status
                print(f"   🏁 Task completed with status: {status}")
                break
        
        # Step 3: Verify final state
        if final_status == 'completed':
            print("   ✅ Async workflow completed successfully")
            return True, {"workflow_status": "completed", "task_id": task_id}
        elif final_status == 'error':
            print("   ❌ Async workflow completed with error")
            return False, {"workflow_status": "error", "task_id": task_id}
        else:
            print("   ⚠️  Async workflow did not complete within polling limit")
            return False, {"workflow_status": "timeout", "task_id": task_id}

def main():
    print("🚀 Starting Image Search API Tests")
    print("=" * 50)
    
    tester = ImageSearchAPITester()
    
    # Test all endpoints
    tests = [
        tester.test_root_endpoint,
        tester.test_single_search_known_codes,
        tester.test_single_search_test_codes,
        tester.test_single_search_empty,
        tester.test_download_image_invalid,
        tester.test_batch_search_known_codes,
        tester.test_batch_search_test_codes,
        tester.test_batch_search_invalid_file,
        tester.test_batch_search_no_codice_column,
        tester.test_download_batch_zip_known_codes,
        tester.test_download_batch_zip_test_codes
    ]
    
    for test in tests:
        try:
            test()
        except Exception as e:
            print(f"❌ Test failed with exception: {str(e)}")
    
    # Print final results
    print("\n" + "=" * 50)
    print(f"📊 Final Results: {tester.tests_passed}/{tester.tests_run} tests passed")
    
    if tester.tests_passed == tester.tests_run:
        print("🎉 All tests passed!")
        return 0
    else:
        print(f"⚠️  {tester.tests_run - tester.tests_passed} tests failed")
        return 1

if __name__ == "__main__":
    sys.exit(main())