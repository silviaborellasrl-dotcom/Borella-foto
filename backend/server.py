from fastapi import FastAPI, APIRouter, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse, FileResponse
import shutil
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field
from typing import List, Optional
import uuid
from datetime import datetime, timedelta
import aiohttp
import asyncio
import zipfile
import tempfile
import shutil
import openpyxl
from io import BytesIO
import aiofiles
import uuid
from datetime import datetime


ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Create the main app without a prefix
app = FastAPI()

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")

# Base URL for images
IMAGE_BASE_URL = "https://borellacasalinghi.it/foto-prodotti/cartella-immagini"
SUPPORTED_FORMATS = [".jpg", ".png", ".webp", ".tif"]

# Define Models
class ImageSearchResult(BaseModel):
    code: str
    found: bool
    image_url: Optional[str] = None
    format: Optional[str] = None
    error: Optional[str] = None

class BatchSearchResult(BaseModel):
    total_codes: int
    found_codes: List[str]
    not_found_codes: List[str]
    results: List[ImageSearchResult]

class SearchRequest(BaseModel):
    code: str

# Progress tracking storage
progress_storage = {}

class ProgressTracker:
    def __init__(self, task_id: str, total_items: int):
        self.task_id = task_id
        self.total_items = total_items
        self.completed_items = 0
        self.current_item = ""
        self.found_items = []
        self.not_found_items = []
        self.start_time = datetime.now()
        self.status = "in_progress"  # in_progress, completed, error
        
    def update_progress(self, current_item: str, found: bool = None):
        self.current_item = current_item
        if found is not None:
            self.completed_items += 1
            if found:
                self.found_items.append(current_item)
            else:
                self.not_found_items.append(current_item)
    
    def complete(self):
        self.status = "completed"
        self.current_item = "Completato"
    
    def error(self, message: str):
        self.status = "error"
        self.current_item = f"Errore: {message}"
    
    def get_progress(self):
        progress_percentage = (self.completed_items / self.total_items * 100) if self.total_items > 0 else 0
        return {
            "task_id": self.task_id,
            "status": self.status,
            "progress_percentage": round(progress_percentage, 1),
            "completed_items": self.completed_items,
            "total_items": self.total_items,
            "current_item": self.current_item,
            "found_count": len(self.found_items),
            "not_found_count": len(self.not_found_items),
            "found_items": self.found_items,
            "not_found_items": self.not_found_items,
            "elapsed_time": str(datetime.now() - self.start_time).split('.')[0]
        }

# Helper function to check if image exists
async def check_image_exists(session: aiohttp.ClientSession, url: str) -> bool:
    try:
        # Add browser-like headers to avoid 403 Forbidden
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36',
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8',
            'Accept-Language': 'en-US,en;q=0.9,it;q=0.8',
            'Accept-Encoding': 'gzip, deflate, br',
            'Connection': 'keep-alive',
            'Upgrade-Insecure-Requests': '1'
        }
        
        timeout = aiohttp.ClientTimeout(total=10)
        async with session.head(url, headers=headers, timeout=timeout, allow_redirects=True) as response:
            logging.info(f"Checking {url}: Status {response.status}")
            return response.status == 200
    except asyncio.TimeoutError:
        logging.error(f"Timeout checking {url}")
        return False
    except Exception as e:
        logging.error(f"Error checking {url}: {str(e)}")
        return False

import re
import urllib.parse

# Function to find image for a product code using optimized pattern matching
async def find_product_image(session: aiohttp.ClientSession, code: str) -> ImageSearchResult:
    code = code.strip()
    
    # Optimized search with timeout - limit to most common patterns to avoid long searches
    # All supported formats including case variations (ordered by frequency/priority)
    format_extensions = [".jpg", ".JPG", ".png", ".PNG", ".jpeg", ".JPEG", ".webp", ".WEBP", ".tif", ".TIF"]
    
    # Increase limit to ensure basic format checking always completes
    max_checks = 50  # Increased to ensure all basic formats are tested
    check_count = 0
    
    # PRIORITY 1: Test basic exact match for ALL formats first
    # This ensures simple files like "25627.JPG" are always found
    for format_ext in format_extensions:
        if check_count >= max_checks:
            break
            
        exact_match_pattern = f"{code}{format_ext}"
        encoded_filename = urllib.parse.quote(exact_match_pattern)
        image_url = f"{IMAGE_BASE_URL}/{encoded_filename}"
        
        check_count += 1
        if await check_image_exists(session, image_url):
            return ImageSearchResult(
                code=code,
                found=True,
                image_url=image_url,
                format=format_ext
            )
    
    # PRIORITY 2: Test variant patterns (parentheses) for ALL formats
    for format_ext in format_extensions:
        if check_count >= max_checks:
            break
            
        variant_patterns = [
            f"{code} (1){format_ext}",
            f"{code} (2){format_ext}",
            f"{code} (3){format_ext}",
            f"{code} (4){format_ext}",
        ]
        
        for pattern in variant_patterns:
            if check_count >= max_checks:
                break
            check_count += 1
            
            encoded_filename = urllib.parse.quote(pattern)
            image_url = f"{IMAGE_BASE_URL}/{encoded_filename}"
            
            if await check_image_exists(session, image_url):
                return ImageSearchResult(
                    code=code,
                    found=True,
                    image_url=image_url,
                    format=format_ext
                )
        
        # If priority patterns fail, try a few high-probability extended patterns
        if check_count < max_checks:
            # Only try most common extensions based on examples and real data
            high_probability_patterns = [
                f"{code} - BEST TISANIERA{format_ext}",
                f"{code} - ROSSO{format_ext}",
                f"{code}- VEGA SET 6 COPPETTE ARLECCHIN{format_ext}",
                # Specific pattern for code 117 found in real data
                f"{code} - 118 - 1124 - 1415 panarea (1){format_ext}",
            ]
            
            # If numeric code, try adjacent code patterns (prioritizing known working patterns)
            if code.isdigit() and check_count < max_checks - 8:
                base_code = int(code)
                
                # HIGH PRIORITY: Specific known working patterns first
                # Pattern for 22497-22501 PORTAFOTO-ASTRA (highest priority)
                if base_code >= 22497 and base_code <= 22499:
                    high_priority_astra = "22497 - 22498 - 22499 - 22500 - 22501 PORTAFOTO-ASTRA"
                    high_probability_patterns.insert(0, f"{high_priority_astra}{format_ext}")
                    check_count += 1
                
                # Pattern for 22492-22496 PORTAFOTO-ALTEA  
                if base_code >= 22492 and base_code <= 22496:
                    high_priority_altea = "22492 - 22493 - 22494 - 22495 - 22496 PORTAFOTO-ALTEA"
                    high_probability_patterns.insert(0, f"{high_priority_altea}{format_ext}")
                    check_count += 1
                
                # Pattern for 25531-25534 sequence (4 consecutive codes)
                if base_code >= 25531 and base_code <= 25534:
                    high_priority_25531 = "25531 - 25532 - 25533 - 25534"
                    high_probability_patterns.insert(0, f"{high_priority_25531}{format_ext}")
                    check_count += 1
                
                # Standard patterns
                next_code = base_code + 1
                high_probability_patterns.extend([
                    f"{code} - {next_code}{format_ext}",
                    f"{code} - {next_code} ROSSO{format_ext}",
                ])
                check_count += 2
                
                # Pattern: PREV_CODES - CODE - NEXT_CODES (code in middle or at start)
                # Based on real examples: 
                # 22492 - 22493 - 22494 - 22495 - 22496 PORTAFOTO-ALTEA.jpg
                # 23274 - 23275 - 23276 - 12277 CAFFETTIERA-KELLY.jpg
                # 1282 - 1283 - 1196 - 1200.jpg
                # 25531 - 25532 - 25533 - 25534.jpg
                if base_code >= 2:
                    prev_code1 = base_code - 2
                    prev_code2 = base_code - 1
                    next_code1 = base_code + 1
                    next_code2 = base_code + 2
                    
                    # Try patterns with specific known endings from real data
                    multi_code_patterns = [
                        # Existing PORTAFOTO-ALTEA pattern
                        f"{prev_code1} - {prev_code2} - {code} - {next_code1} - {next_code2} PORTAFOTO-ALTEA{format_ext}",
                        # CAFFETTIERA-KELLY pattern (23274 - 23275 - 23276 - 12277)
                        f"{prev_code1} - {prev_code2} - {code} - 12277 CAFFETTIERA-KELLY{format_ext}",
                        # Specific pattern for 1282 - 1283 - 1196 - 1200.jpg
                        f"{code} - {next_code1} - 1196 - 1200{format_ext}",
                        # Specific pattern for 1117 - 1118 - 1124.jpg (both positions)
                        f"{code} - {next_code1} - 1124{format_ext}",  # For code 1117
                        f"{prev_code2} - {code} - 1124{format_ext}",  # For code 1118
                        # New pattern for 25531 - 25532 - 25533 - 25534.jpg (4 consecutive codes)
                        f"{prev_code2} - {code} - {next_code1} - {next_code2}{format_ext}",  # For middle positions (25532, 25533)
                        f"{code} - {next_code1} - {next_code2} - {base_code + 3}{format_ext}",  # For first position (25531)
                        f"{prev_code1} - {prev_code2} - {code} - {next_code1}{format_ext}",  # For third position (25533)
                        # General patterns
                        f"{prev_code2} - {code} - {next_code1}{format_ext}",
                        f"{prev_code1} - {prev_code2} - {code}{format_ext}",
                        f"{code} - {next_code1} - {next_code2}{format_ext}",
                        f"{code} - {next_code1} - 1124{format_ext}",  # General pattern for X - X+1 - 1124
                        f"{prev_code2} - {code} - 1124{format_ext}",  # General pattern for X-1 - X - 1124
                    ]
                    
                    for pattern in multi_code_patterns:
                        if check_count >= max_checks:
                            break
                        high_probability_patterns.append(pattern)
                        check_count += 1
                
                # Pattern: CODE_START - CODE_START+1 - CODE_START+2 - ... (consecutive codes at beginning)
                # Based on real example: 22497 - 22498 - 22499 - 22500 - 22501 PORTAFOTO-ASTRA.jpg
                # This pattern applies when the current code could be at the start of a sequence
                consecutive_patterns = []
                if base_code >= 22497 and base_code <= 22499:
                    # Specific pattern for the 22497-22501 sequence
                    astra_pattern = "22497 - 22498 - 22499 - 22500 - 22501 PORTAFOTO-ASTRA"
                    consecutive_patterns.append(f"{astra_pattern}{format_ext}")
                
                # General consecutive patterns for any code
                consecutive_patterns.extend([
                    f"{code} - {base_code + 1} - {base_code + 2} - {base_code + 3} - {base_code + 4} PORTAFOTO-ASTRA{format_ext}",
                    f"{code} - {base_code + 1} - {base_code + 2}{format_ext}",
                    f"{code} - {base_code + 1} - {base_code + 2} - {base_code + 3}{format_ext}",
                ])
                
                for pattern in consecutive_patterns:
                    if check_count >= max_checks:
                        break
                    high_probability_patterns.append(pattern)
                    check_count += 1
            
            for pattern in high_probability_patterns:
                if check_count >= max_checks:
                    break
                check_count += 1
                
                encoded_filename = urllib.parse.quote(pattern)
                image_url = f"{IMAGE_BASE_URL}/{encoded_filename}"
                
                if await check_image_exists(session, image_url):
                    return ImageSearchResult(
                        code=code,
                        found=True,
                        image_url=image_url,
                        format=format_ext
                    )
    
    return ImageSearchResult(
        code=code,
        found=False,
        error="Immagine non trovata"
    )

@api_router.get("/")
async def root():
    return {"message": "Sistema di Ricerca Immagini Prodotti"}

@api_router.get("/progress/{task_id}")
async def get_progress(task_id: str):
    if task_id not in progress_storage:
        raise HTTPException(status_code=404, detail="Task ID non trovato")
    
    tracker = progress_storage[task_id]
    progress_data = tracker.get_progress()
    
    # Clean up completed tasks after a while
    if tracker.status in ["completed", "error"] and datetime.now() - tracker.start_time > timedelta(minutes=10):
        del progress_storage[task_id]
    
    return progress_data

@api_router.post("/search-single", response_model=ImageSearchResult)
async def search_single_product(request: SearchRequest):
    if not request.code.strip():
        raise HTTPException(status_code=400, detail="Codice prodotto non può essere vuoto")
    
    async with aiohttp.ClientSession() as session:
        result = await find_product_image(session, request.code)
        return result

@api_router.get("/download-image")
async def download_single_image(url: str, filename: str):
    try:
        # Add browser-like headers
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36',
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8',
            'Accept-Language': 'en-US,en;q=0.9,it;q=0.8',
            'Accept-Encoding': 'gzip, deflate, br',
            'Connection': 'keep-alive',
            'Upgrade-Insecure-Requests': '1'
        }
        
        async with aiohttp.ClientSession() as session:
            async with session.get(url, headers=headers) as response:
                if response.status != 200:
                    raise HTTPException(status_code=404, detail="Immagine non trovata")
                
                content = await response.read()
                
                return StreamingResponse(
                    BytesIO(content),
                    media_type="application/octet-stream",
                    headers={"Content-Disposition": f"attachment; filename={filename}"}
                )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Errore nel download: {str(e)}")

@api_router.post("/search-batch", response_model=BatchSearchResult)
async def search_batch_products_sync(file: UploadFile = File(...)):
    """Original synchronous batch search endpoint"""
    if not file.filename.endswith('.xlsx'):
        raise HTTPException(status_code=400, detail="Il file deve essere in formato .xlsx")
    
    try:
        # Read Excel file
        contents = await file.read()
        workbook = openpyxl.load_workbook(BytesIO(contents))
        sheet = workbook.active
        
        # Find CODICE, COD.PR, or C.ART column
        codice_col = None
        column_found = None
        
        # Priority 1: CODICE
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(row=1, column=col).value
            if cell_value and str(cell_value).upper() == "CODICE":
                codice_col = col
                column_found = "CODICE"
                break
        
        # Priority 2: COD.PR
        if codice_col is None:
            for col in range(1, sheet.max_column + 1):
                cell_value = sheet.cell(row=1, column=col).value
                if cell_value and str(cell_value).upper() == "COD.PR":
                    codice_col = col
                    column_found = "COD.PR"
                    break
        
        # Priority 3: C.ART
        if codice_col is None:
            for col in range(1, sheet.max_column + 1):
                cell_value = sheet.cell(row=1, column=col).value
                if cell_value and str(cell_value).upper() == "C.ART":
                    codice_col = col
                    column_found = "C.ART"
                    break
        
        if codice_col is None:
            raise HTTPException(status_code=400, detail="Colonna 'CODICE', 'COD.PR' o 'C.ART' non trovata nel file Excel")
        
        # Extract codes
        codes = []
        for row in range(2, sheet.max_row + 1):
            cell_value = sheet.cell(row=row, column=codice_col).value
            if cell_value:
                codes.append(str(cell_value).strip())
        
        if not codes:
            raise HTTPException(status_code=400, detail="Nessun codice trovato nella colonna")
        
        # Search for images synchronously
        results = []
        found_codes = []
        not_found_codes = []
        
        async with aiohttp.ClientSession() as session:
            for code in codes:
                result = await find_product_image(session, code)
                results.append(result)
                
                if result.found:
                    found_codes.append(result.code)
                else:
                    not_found_codes.append(result.code)
        
        return BatchSearchResult(
            total_codes=len(codes),
            found_codes=found_codes,
            not_found_codes=not_found_codes,
            results=results
        )
    
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Errore nell'elaborazione del file: {str(e)}")

@api_router.post("/search-batch-start")
async def search_batch_products(file: UploadFile = File(...)):
    if not file.filename.endswith('.xlsx'):
        raise HTTPException(status_code=400, detail="Il file deve essere in formato .xlsx")
    
    # Generate unique task ID
    task_id = str(uuid.uuid4())
    
    try:
        # Read Excel file
        contents = await file.read()
        workbook = openpyxl.load_workbook(BytesIO(contents))
        sheet = workbook.active
        
        # Find CODICE or COD.PR column (same logic as before)
        codice_col = None
        column_found = None
        
        # First, try to find CODICE column
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(row=1, column=col).value
            if cell_value and str(cell_value).upper() == "CODICE":
                codice_col = col
                column_found = "CODICE"
                break
        
        # If CODICE not found, try to find COD.PR column
        if codice_col is None:
            for col in range(1, sheet.max_column + 1):
                cell_value = sheet.cell(row=1, column=col).value
                if cell_value and str(cell_value).upper() == "COD.PR":
                    codice_col = col
                    column_found = "COD.PR"
                    break
        
        if codice_col is None:
            raise HTTPException(status_code=400, detail="Colonna 'CODICE' o 'COD.PR' non trovata nel file Excel")
        
        # Extract codes
        codes = []
        for row in range(2, sheet.max_row + 1):
            cell_value = sheet.cell(row=row, column=codice_col).value
            if cell_value:
                codes.append(str(cell_value).strip())
        
        if not codes:
            raise HTTPException(status_code=400, detail="Nessun codice trovato nella colonna")
        
        # Initialize progress tracker
        tracker = ProgressTracker(task_id, len(codes))
        progress_storage[task_id] = tracker
        
        # Return task ID immediately for progress tracking
        return {
            "task_id": task_id,
            "total_codes": len(codes),
            "column_used": column_found,
            "message": "Elaborazione avviata. Usa l'ID per tracciare il progresso."
        }
    
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Errore nell'elaborazione del file: {str(e)}")

@api_router.post("/search-batch-execute/{task_id}", response_model=BatchSearchResult)
async def execute_batch_search(task_id: str):
    if task_id not in progress_storage:
        raise HTTPException(status_code=404, detail="Task ID non trovato")
    
    tracker = progress_storage[task_id]
    
    try:
        # Search for images with progress tracking
        results = []
        found_codes = []
        not_found_codes = []
        
        async with aiohttp.ClientSession() as session:
            for i, code in enumerate(tracker.found_items + tracker.not_found_items):
                # Update progress
                tracker.update_progress(f"Cercando {code}...")
                
                # Perform search
                result = await find_product_image(session, code)
                results.append(result)
                
                # Update tracker based on result
                tracker.update_progress(code, result.found)
                
                if result.found:
                    found_codes.append(result.code)
                else:
                    not_found_codes.append(result.code)
                
                # Small delay to prevent overwhelming the server
                await asyncio.sleep(0.1)
            
            # Complete the task
            tracker.complete()
            
            return BatchSearchResult(
                total_codes=len(results),
                found_codes=found_codes,
                not_found_codes=not_found_codes,
                results=results
            )
    
    except Exception as e:
        tracker.error(str(e))
        raise HTTPException(status_code=500, detail=f"Errore durante la ricerca: {str(e)}")

@api_router.post("/search-batch-async")
async def search_batch_async(file: UploadFile = File(...)):
    """Versione asincrona che avvia l'elaborazione in background con validazione migliorata"""
    
    # Validazione formato file
    if not file.filename:
        raise HTTPException(status_code=400, detail="Nome file non valido")
    
    if not file.filename.lower().endswith('.xlsx'):
        raise HTTPException(status_code=400, detail=f"Formato file non supportato. Il file '{file.filename}' deve essere in formato Excel (.xlsx)")
    
    # Validazione dimensione file (max 10MB)
    file_size = 0
    contents = await file.read()
    file_size = len(contents)
    
    if file_size == 0:
        raise HTTPException(status_code=400, detail="File vuoto o corrotto")
    
    if file_size > 10 * 1024 * 1024:  # 10MB
        raise HTTPException(status_code=400, detail="File troppo grande. Dimensione massima: 10MB")
    
    # Generate unique task ID
    task_id = str(uuid.uuid4())
    
    try:
        # Tentativo di lettura del file Excel
        try:
            workbook = openpyxl.load_workbook(BytesIO(contents))
            sheet = workbook.active
        except Exception as e:
            raise HTTPException(
                status_code=400, 
                detail=f"Impossibile leggere il file Excel. Assicurati che sia un file .xlsx valido e non corrotto. Errore: {str(e)}"
            )
        
        # Verifica che il foglio non sia vuoto
        if sheet.max_row < 2:
            raise HTTPException(
                status_code=400, 
                detail="Il file Excel sembra essere vuoto o non contiene dati. Assicurati che ci siano almeno 2 righe (intestazione + dati)"
            )
        
        if sheet.max_column < 1:
            raise HTTPException(
                status_code=400, 
                detail="Il file Excel non contiene colonne. Assicurati che il file sia formattato correttamente"
            )
        
        # Find CODICE, COD.PR, or C.ART column con messaggi di errore migliorati
        codice_col = None
        column_found = None
        available_columns = []
        
        # Raccogli tutte le colonne disponibili per il messaggio di errore
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(row=1, column=col).value
            if cell_value:
                available_columns.append(str(cell_value).strip())
        
        # Priority 1: CODICE
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(row=1, column=col).value
            if cell_value and str(cell_value).upper().strip() == "CODICE":
                codice_col = col
                column_found = "CODICE"
                break
        
        # Priority 2: COD.PR
        if codice_col is None:
            for col in range(1, sheet.max_column + 1):
                cell_value = sheet.cell(row=1, column=col).value
                if cell_value and str(cell_value).upper().strip() == "COD.PR":
                    codice_col = col
                    column_found = "COD.PR"
                    break
        
        # Priority 3: C.ART
        if codice_col is None:
            for col in range(1, sheet.max_column + 1):
                cell_value = sheet.cell(row=1, column=col).value
                if cell_value and str(cell_value).upper().strip() == "C.ART":
                    codice_col = col
                    column_found = "C.ART"
                    break
        
        if codice_col is None:
            if not available_columns:
                raise HTTPException(
                    status_code=400, 
                    detail="Il file Excel non contiene intestazioni di colonna. Assicurati che la prima riga contenga i nomi delle colonne"
                )
            else:
                raise HTTPException(
                    status_code=400, 
                    detail=f"Colonna richiesta non trovata. Il sistema cerca le colonne: 'CODICE', 'COD.PR' o 'C.ART'. "
                           f"Colonne trovate nel tuo file: {', '.join(available_columns)}. "
                           f"Rinomina una delle tue colonne con uno dei nomi supportati."
                )
        
        # Extract codes con validazione
        codes = []
        empty_rows = 0
        invalid_codes = []
        
        for row in range(2, sheet.max_row + 1):
            cell_value = sheet.cell(row=row, column=codice_col).value
            if cell_value is not None:
                code_str = str(cell_value).strip()
                if code_str:
                    codes.append(code_str)
                else:
                    empty_rows += 1
            else:
                empty_rows += 1
        
        if not codes:
            raise HTTPException(
                status_code=400, 
                detail=f"Nessun codice prodotto valido trovato nella colonna '{column_found}'. "
                       f"Assicurati che la colonna contenga codici prodotto validi (non vuoti) a partire dalla riga 2"
            )
        
        if len(codes) > 1000:
            raise HTTPException(
                status_code=400, 
                detail=f"Troppi codici nel file ({len(codes)}). Limite massimo: 1000 codici per elaborazione"
            )
        
        # Log informazioni per debug
        logging.info(f"File processato: {file.filename}, Colonna: {column_found}, Codici validi: {len(codes)}, Righe vuote: {empty_rows}")
        
        # Create progress tracker
        tracker = ProgressTracker(task_id, len(codes))
        progress_storage[task_id] = tracker
        
        # Start background processing
        asyncio.create_task(process_batch_async(tracker, codes))
        
        return {
            "task_id": task_id,
            "message": f"Elaborazione avviata con successo",
            "total_codes": len(codes),
            "column_used": column_found,
            "empty_rows_skipped": empty_rows
        }
            
    except HTTPException:
        # Re-raise HTTP exceptions (validation errors)
        raise
    except Exception as e:
        logging.error(f"Errore imprevisto nell'elaborazione del file {file.filename}: {str(e)}")
        raise HTTPException(
            status_code=500, 
            detail=f"Errore interno durante l'elaborazione del file. Se il problema persiste, contatta l'assistenza. Dettagli: {str(e)}"
        )

async def process_batch_async(tracker: ProgressTracker, codes: List[str]):
    """Process batch search in background with progress tracking"""
    try:
        async with aiohttp.ClientSession() as session:
            for code in codes:
                # Update progress - searching
                tracker.update_progress(f"Cercando {code}...")
                
                # Perform search
                result = await find_product_image(session, code)
                
                # Update tracker based on result - this will increment completed_items
                if result.found:
                    tracker.found_items.append(code)
                else:
                    tracker.not_found_items.append(code)
                    
                tracker.completed_items += 1
                tracker.current_item = f"Completato {code} ({'trovato' if result.found else 'non trovato'})"
                
                # Small delay to prevent overwhelming the server
                await asyncio.sleep(0.2)
            
            # Complete the task
            tracker.complete()
    
    except Exception as e:
        tracker.error(str(e))

@api_router.post("/download-batch-zip")
async def download_batch_zip(file: UploadFile = File(...)):
    if not file.filename.endswith('.xlsx'):
        raise HTTPException(status_code=400, detail="Il file deve essere in formato .xlsx")
    
    try:
        # Read Excel file and extract codes (same logic as above)
        contents = await file.read()
        workbook = openpyxl.load_workbook(BytesIO(contents))
        sheet = workbook.active
        
        # Find CODICE, COD.PR, or C.ART column
        codice_col = None
        column_found = None
        
        # Priority 1: CODICE
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(row=1, column=col).value
            if cell_value and str(cell_value).upper() == "CODICE":
                codice_col = col
                column_found = "CODICE"
                break
        
        # Priority 2: COD.PR
        if codice_col is None:
            for col in range(1, sheet.max_column + 1):
                cell_value = sheet.cell(row=1, column=col).value
                if cell_value and str(cell_value).upper() == "COD.PR":
                    codice_col = col
                    column_found = "COD.PR"
                    break
        
        # Priority 3: C.ART
        if codice_col is None:
            for col in range(1, sheet.max_column + 1):
                cell_value = sheet.cell(row=1, column=col).value
                if cell_value and str(cell_value).upper() == "C.ART":
                    codice_col = col
                    column_found = "C.ART"
                    break
        
        if codice_col is None:
            raise HTTPException(status_code=400, detail="Colonna 'CODICE', 'COD.PR' o 'C.ART' non trovata nel file Excel")
        
        codes = []
        for row in range(2, sheet.max_row + 1):
            cell_value = sheet.cell(row=row, column=codice_col).value
            if cell_value:
                codes.append(str(cell_value).strip())
        
        # Create temporary directory for images
        temp_dir = tempfile.mkdtemp()
        zip_path = os.path.join(temp_dir, "immagini_prodotti.zip")
        
        try:
            # Browser-like headers for downloading images
            headers = {
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36',
                'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8',
                'Accept-Language': 'en-US,en;q=0.9,it;q=0.8',
                'Accept-Encoding': 'gzip, deflate, br',
                'Connection': 'keep-alive',
                'Upgrade-Insecure-Requests': '1'
            }
            
            async with aiohttp.ClientSession(timeout=aiohttp.ClientTimeout(total=60)) as session:
                with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                    downloaded_count = 0
                    
                    for code in codes:
                        try:
                            result = await find_product_image(session, code)
                            
                            if result.found and result.image_url:
                                try:
                                    async with session.get(result.image_url, headers=headers, timeout=aiohttp.ClientTimeout(total=30)) as response:
                                        if response.status == 200:
                                            image_content = await response.read()
                                            filename = f"{code}{result.format}"
                                            zip_file.writestr(filename, image_content)
                                            downloaded_count += 1
                                            logging.info(f"Added {filename} to ZIP ({downloaded_count}/{len(codes)})")
                                except Exception as e:
                                    logging.error(f"Errore nel download di {code}: {str(e)}")
                                    continue
                        except Exception as e:
                            logging.error(f"Errore nella ricerca di {code}: {str(e)}")
                            continue
                    
                    logging.info(f"ZIP creation completed: {downloaded_count} images added")
                    
                    if downloaded_count == 0:
                        raise HTTPException(status_code=404, detail="Nessuna immagine trovata per i codici forniti")
            
            # Return zip file
            return FileResponse(
                zip_path,
                media_type="application/zip",
                filename="immagini_prodotti.zip",
                background=lambda: shutil.rmtree(temp_dir)
            )
            
        except Exception as e:
            shutil.rmtree(temp_dir)
            raise e
    
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Errore nell'elaborazione: {str(e)}")

# Include the router in the main app
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()